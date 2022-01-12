from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
gradesheet_workbook = load_workbook("files\\MAT110 FALL 2021 Gradesheet.xlsx")
mid_wb = load_workbook("files\\MAT110 Midterm Exam Fall21 (Responses).xlsx")["Form Responses 1"]
quiz_wb = load_workbook("files\\Quiz Marks.xlsx")["Worksheet"]
final_date = load_workbook("files\Date-of-Final-Exam-_Responses_-Form-Responses-1.xlsx")["Worksheet"]

count=0
output = ""
information = Workbook()
# forbot = Workbook()

# bot = forbot.create_sheet("Quiz 1")
# bot["A"+str(1)] = "Student ID"
# bot["B"+str(1)] = "Name"
# bot["C"+str(1)] = "G-Suite"
# bot["D"+str(1)] = "BUX Username"
# bot["E"+str(1)] = "Quiz Mark"
# botcount = 1

totalcount=0
absentcount = 0
for gradesheet_ws in gradesheet_workbook.worksheets:
    infos = information.create_sheet(gradesheet_ws.title)
    infos["A"+str(1)] = "Student ID"
    infos["B"+str(1)] = "Name"
    infos["C"+str(1)] = "G-Suite"
    infos["D"+str(1)] = "BUX Username"
    infos["E"+str(1)] = "Quiz 1"
    infos["F"+str(1)] = "Quiz 2"
    infos["G"+str(1)] = "Quiz 3"
    infos["H"+str(1)] = "Quiz Average"


    count_sec=0
    loss_count_sec=0

    absent=[]
    for row in range(4, 60):
        
        got = None
        id_char  = get_column_letter(2)+str(row)
        name_char  = get_column_letter(3)+str(row)
        id = str(gradesheet_ws[id_char].value)
        if '.' in id: id=id[:-2]
        name = str(gradesheet_ws[name_char].value)
        if id!="None":
            got=False
        if id=="None":
            break
        # botcount+=1
        infos["A"+str(row-2)] = id
        infos["B"+str(row-2)] = name
        # bot["A"+str(botcount)] = id
        # bot["B"+str(botcount)] = name
        # for mid_row in range(2,845):
        for final_row in range(2, 1326):
            final_id_char = "D"+str(final_row)
            # mid_id_char = "F"+str(mid_row)
            # mid_id = str(mid_wb[mid_id_char].value)
            final_id = str(final_date[final_id_char].value)
            # if '.' in mid_id: mid_id=mid_id[:-2]
            if '.' in final_id: final_id=final_id[:-2]
            # print(final_id)
            # if mid_id in id:
            if final_id in id:
                got = True
                count_sec+=1
                final_email_char = "B"+str(final_row)
                # mid_email_char = "B"+str(mid_row)
                # mid_email2_char = get_column_letter(3)+str(mid_row)
                # mid_bux_char = get_column_letter(4)+str(mid_row)
                final_email = str(final_date[final_email_char].value)
                # mid_email = str(mid_wb[mid_email_char].value)
                # mid_email2 = str(mid_wb[mid_email2_char].value)

                for quiz_row in range(2, 1110):
                    quiz_bux_char = "C"+str(quiz_row)
                    quiz_email_char = get_column_letter(2)+str(quiz_row)
                    quiz_marks_char = "AP"+str(quiz_row)
                    quiz_marks_char2 = "AQ"+str(quiz_row)
                    quiz_marks_char3 = "AR"+str(quiz_row)
                    quiz_marks_char_avg = "AS"+str(quiz_row)
                    quiz_email = str(quiz_wb[quiz_email_char].value)
                    quiz_marks = quiz_wb[quiz_marks_char].value
                    quiz_marks2 = quiz_wb[quiz_marks_char2].value
                    quiz_marks3 = quiz_wb[quiz_marks_char3].value
                    quiz_marks_avg = quiz_wb[quiz_marks_char_avg].value
                    
                    quiz_bux = quiz_wb[quiz_bux_char].value
                    if quiz_marks!="Not Attempted": quiz_marks = int(float(quiz_marks)*25) 
                    else: quiz_marks=0
                    if quiz_marks2!="Not Attempted": quiz_marks2 = int(float(quiz_marks2)*25) 
                    else: quiz_marks2=0
                    if quiz_marks3!="Not Attempted": quiz_marks3 = int(float(quiz_marks3)*25) 
                    else: quiz_marks3=0
                    quiz_marks_avg = int(float(quiz_marks_avg)*25) 
                    # if quiz_email == mid_email or mid_email2==quiz_email:
                    if quiz_email == final_email:

                        count+=1
                        gradesheet_ws["P"+str(row)] = quiz_marks
                        gradesheet_ws["Q"+str(row)] = quiz_marks2
                        gradesheet_ws["R"+str(row)] = quiz_marks3
                        gradesheet_ws["S"+str(row)] = quiz_marks_avg

                        # infos["C"+str(row-2)] = mid_email
                        infos["C"+str(row-2)] = final_email
                        infos["D"+str(row-2)] = quiz_bux
                        infos["E"+str(row-2)] = quiz_marks
                        infos["F"+str(row-2)] = quiz_marks2
                        infos["G"+str(row-2)] = quiz_marks3
                        infos["H"+str(row-2)] = quiz_marks_avg

                        totalcount+=1
                        # bot["C"+str(botcount)] = mid_email
                        # bot["D"+str(botcount)] = quiz_bux
                        # bot["E"+str(botcount)] = quiz_marks
                        break
                break
        if got== False:
            loss_count_sec+=1
            absent.append("ID: "+id+"  Name: "+name)
    print("\n\n\n"+gradesheet_ws.title+":",count_sec,"students out of",loss_count_sec+count_sec,"students are listed\n")
    output+="\n\n\n"+gradesheet_ws.title+": "+str(count_sec)+" students out of "+str(loss_count_sec+count_sec)+" students are listed.\n\n"
    print("Not Listed: \n--------------------------\n\n")
    for abse in absent:
        print(abse)
        absentcount+=1
        output+="\n"+abse
    information.save("information.xlsx")
    gradesheet_workbook.save("new gradesheet.xlsx")
    # forbot.save("./mat110_discord_bot/MAT110API.xlsx")
    output+="\n"
print("\nTotal Listed:",totalcount)
output+="\n\nTotal Listed: "+str(totalcount)
output+="\n\nNot Listed: "+str(absentcount)
print("\nNot Listed: "+str(absentcount))

file = open("output.txt","w")
file.write(output)